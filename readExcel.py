from openpyxl import load_workbook
class OperateExcel(object):
    rows = None
    cols = None
    table = None
    excel = None
    def __init__(self,path,sheetname):
        # 打开文件：
        self.excel = load_workbook(path)
        # 获取sheet：
        self.table = self.excel[sheetname]  # 通过表名获取
        # 获取行数和列数：
        self.rows = self.table.max_row  # 获取行数
        self.cols = self.table.max_column  # 获取列数

    # 设置单元格值
    def set_value(self,row,col,value):
        self.table.cell(row, col).value = value

    # 返回单元格值
    def get_value(self,row,col):
        return self.table.cell(row, col).value

    # 保存文件
    def save(self,path):
        self.excel.save(path)




